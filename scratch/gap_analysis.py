import openpyxl
import json
import re

# We will load the populator mapping from the python file directly
# Or just copy the mapping here to compare
POPULATOR_MAPPING = {
    "share_capital": 204,
    "reserves_and_surplus": 205,
    "money_received_share_warrants": 206,
    "share_application_money": 208,
    "long_term_borrowings": 212,
    "deferred_tax_liabilities": 213,
    "other_long_term_liabilities": 214,
    "long_term_provisions": 215,
    "short_term_borrowings": 218,
    "trade_payables_msme": 220,
    "trade_payables_others": 221,
    "other_current_liabilities": 222,
    "short_term_provisions": 223,
    "tangible_assets": 229,
    "intangible_assets": 231,
    "capital_wip": 232,
    "intangible_assets_under_dev": 233,
    "non_current_investments": 234,
    "deferred_tax_assets": 235,
    "long_term_loans_advances": 236,
    "other_non_current_assets": 237,
    "current_investments": 239,
    "inventories": 240,
    "trade_receivables": 241,
    "cash_and_bank_balances": 242,
    "short_term_loans_advances": 243,
    "other_current_assets": 244,
    "rev_sale_goods_mfg": 449,
    "rev_sale_goods_traded": 450,
    "rev_sale_services": 451,
    "oi_dividend": 457,
    "oi_interest": 458,
    "oi_net_gain_investments": 459,
    "oi_other_non_operating": 460,
    "cost_of_materials_consumed": 463,
    "purchases_of_stock_in_trade": 464,
    "changes_in_inventories": 465,
    "employee_benefit_expense": 468,
    "managerial_remuneration": 469,
    "payment_to_auditors": 470,
    "insurance_expenses": 471,
    "power_and_fuel": 472,
    "finance_costs": 473,
    "depreciation_and_amortisation": 474,
    "other_expenses": 475,
    "exceptional_items": 478,
    "extraordinary_items": 480,
    "current_tax": 483,
    "deferred_tax": 484
}

def analyze():
    # Load diagnosis json
    try:
        with open(r"c:\RIYAS\Sharp INtell\SI Filings\aoc4_diagnosis.json", "r", encoding="utf-8") as f:
            diagnosis_data = json.load(f)
    except Exception as e:
        print(f"Error loading diagnosis: {e}")
        diagnosis_data = {}

    # Load Excel
    filepath = r"c:\RIYAS\Sharp INtell\SI Filings\AOC-4_U92410KL2020PTC065216_2021-2022_20260729.xlsx"
    print(f"Inspecting Excel: {filepath}\n")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb["FORM"]
    except Exception as e:
        print(f"Error loading Excel: {e}")
        return

    # To check which row is mapped to which key
    reverse_map = {v: k for k, v in POPULATOR_MAPPING.items()}
    
    # We look for rows that have labels in Column C or B (from row 190 to 550)
    print("--- GAP ANALYSIS REPORT ---")
    print("Format: Row | Excel Label | Mapped Key | In JSON? | CY Value in JSON | PY Value in JSON\n")
    
    for row in range(198, 550):
        # The label is usually in column C, sometimes in column B
        label = str(sheet.cell(row=row, column=3).value or "").strip()
        if not label or label.lower() in ("none", ""):
            label = str(sheet.cell(row=row, column=2).value or "").strip()
            if not label or label.lower() in ("none", ""):
                continue

        # Skip headers
        if "Particulars" in label or label.startswith("I") or label.startswith("II") or label.startswith("III"):
            # Wait, some sections are III Financial parameters etc
            # Let's keep it to see the structure
            pass
            
        clean_label = label.replace("\n", " ")
        if len(clean_label) > 60:
            clean_label = clean_label[:57] + "..."
            
        mapped_key = reverse_map.get(row, "NOT_MAPPED")
        
        in_json = "NO"
        cy_val = ""
        py_val = ""
        
        if mapped_key != "NOT_MAPPED":
            if mapped_key in diagnosis_data:
                in_json = "YES"
                val = diagnosis_data[mapped_key]
                if isinstance(val, dict):
                    cy_val = val.get("current_year", "")
                    py_val = val.get("previous_year", "")
                else:
                    cy_val = val
        else:
            # Let's see if we can find a matching key based on label
            # Just rough
            pass
            
        print(f"Row {row:3d} | {clean_label:<60} | {mapped_key:<28} | JSON: {in_json:<3} | CY: {cy_val:<10} | PY: {py_val:<10}")

if __name__ == "__main__":
    analyze()
