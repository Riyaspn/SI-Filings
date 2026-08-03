"""
Verify extracted JSON payload against CS-completed reference Excel file.
"""
import openpyxl

ref_path = r"C:\Users\RIYAS\Downloads\Copy of AOC-4_U92410KL2020PTC065216_2021-2022_20260728.xlsx"

wb = openpyxl.load_workbook(ref_path, data_only=True)
sheet = wb["FORM"]

# Key field mappings (row, col_cy, col_py, json_key)
field_mappings = [
    # GENERAL INFO
    (22, 2, None, "company_name", "Company Name"),
    (33, 2, None, "nature_of_financial_statements", "Nature of FS"),
    (83, 7, None, "agm_date", "AGM Date"),
    (97, 7, None, "auditor_frn", "Auditor FRN"),
    (98, 7, None, "auditor_name", "Auditor Name"),
    
    # BALANCE SHEET
    (204, 7, 10, "share_capital", "Share Capital"),
    (205, 7, 10, "reserves_and_surplus", "Reserves & Surplus"),
    (212, 7, 10, "long_term_borrowings", "Long-Term Borrowings"),
    (213, 7, 10, "deferred_tax_liabilities", "Deferred Tax Liabilities"),
    (222, 7, 10, "other_current_liabilities", "Other Current Liabilities"),
    (224, 7, 10, "total_equity_and_liabilities", "Total Equity & Liabilities"),
    (229, 7, 10, "tangible_assets", "Property Plant & Equipment (Net BV)"),
    (235, 7, 10, "deferred_tax_assets", "Deferred Tax Assets"),
    (242, 7, 10, "cash_and_bank_balances", "Cash & Bank Balances"),
    (244, 7, 10, "other_current_assets", "Other Current Assets"),
    (245, 7, 10, "total_assets", "Total Assets"),
    
    # BREAK-UP
    (258, 10, 10, "ltb_loans_related", "LTB Loans Related Parties"),
    
    # FINANCIAL PARAMS BS
    (351, 13, 14, "gross_transaction_as_18", "AS-18 Gross Transactions (BS Params)"),
    (371, 13, 14, "net_worth", "Net Worth"),
    (374, 13, 14, "gross_ppe", "Gross PPE"),
    (375, 13, 14, "depreciation_and_amortisation", "Accumulated Depreciation"),
    
    # P&L STATEMENT
    (449, 7, 10, "rev_sale_goods_mfg", "Rev: Goods Mfg"),
    (451, 7, 10, "rev_sale_services", "Rev: Services"),
    (461, 7, 10, "revenue_from_operations", "Total Revenue / Income"),
    (468, 7, 10, "employee_benefit_expense", "Employee Benefit Expense"),
    (470, 7, 10, "payment_to_auditors", "Payment to Auditors"),
    (472, 7, 10, "power_and_fuel", "Power and Fuel"),
    (473, 7, 10, "finance_costs", "Finance Costs"),
    (474, 7, 10, "depreciation_and_amortisation", "Depreciation & Amortisation"),
    (475, 7, 10, "other_expenses", "Other Expenses"),
    (476, 7, 10, "total_expenses", "Total Expenses"),
    (477, 7, 10, "profit_before_exceptional_items", "Profit Before Tax"),
    (484, 7, 10, "deferred_tax", "Deferred Tax Expense"),
    (485, 7, 10, "profit_after_tax", "Profit/(Loss) for the period"),
    (491, 7, 10, "earnings_per_share_basic", "Basic EPS"),
    
    # FINANCIAL PARAMS PL
    (532, 13, 14, "param_rent_paid", "Rent Paid (PL Params)"),
    (534, 13, 14, "gross_transaction_as_18", "AS-18 Gross Transactions (PL Params)"),
    
    # PRINCIPAL PRODUCTS
    (543, 2, None, "pcs_code", "Product Code (4-digit)"),
    (543, 7, None, "pcs_turnover", "Product Turnover"),
    (543, 9, None, "pcs_highest_code", "Highest Product Code (8-digit)"),
    (543, 11, None, "pcs_description", "Product Description"),
    (543, 13, None, "pcs_highest_turnover", "Highest Product Turnover"),
]

# Extract data dictionary from user JSON
import json
json_str = '''
{
  "cin": "U92410KL2020PTC065216",
  "company_name": "VIBESEK VENTURES PRIVATE LIMITED",
  "reporting_unit": "Hundreds",
  "fy_start_date": "2021-04-01",
  "fy_end_date": "2022-03-31",
  "board_meeting_date": "2026-06-26",
  "nature_of_financial_statements": "Adopted Financial statements",
  "auditor_report_date": "2026-06-26",
  "category_of_auditor": "Limited Liability Partnership (LLP)",
  "auditor_frn": "000158S",
  "auditor_name": "C J & Co.LLP",
  "agm_held": "Yes",
  "agm_date": "",
  "agm_due_date": "2022-09-30",
  "share_capital": {"current_year": 10000.0, "previous_year": 10000.0},
  "reserves_and_surplus": {"current_year": -65164.0, "previous_year": -24323.0},
  "long_term_borrowings": {"current_year": 2037264.0, "previous_year": 815620.0},
  "ltb_term_loans_others": {"current_year": 0.0, "previous_year": 0.0},
  "ltb_loans_related": {"current_year": 2037264.0, "previous_year": 815620.0},
  "deferred_tax_liabilities": {"current_year": 8300.0, "previous_year": 0.0},
  "other_current_liabilities": {"current_year": 10000.0, "previous_year": 7500.0},
  "total_equity_and_liabilities": {"current_year": 2000400.0, "previous_year": 808797.0},
  "tangible_assets": {"current_year": 1395912.0, "previous_year": 0.0},
  "gross_ppe": {"current_year": 1422571.0, "previous_year": 0.0},
  "accumulated_depreciation_ppe": {"current_year": 26659.0, "previous_year": 0.0},
  "capital_wip": {"current_year": 0.0, "previous_year": 135226.0},
  "deferred_tax_assets": {"current_year": 0.0, "previous_year": 3926.0},
  "cash_and_bank_balances": {"current_year": 8388.0, "previous_year": 69645.0},
  "other_current_assets": {"current_year": 596100.0, "previous_year": 600000.0},
  "total_assets": {"current_year": 2000400.0, "previous_year": 808797.0},
  "revenue_from_operations": {"current_year": 416194.0, "previous_year": 0.0},
  "rev_sale_goods_mfg": {"current_year": 0.0, "previous_year": 0.0},
  "rev_sale_services": {"current_year": 416194.0, "previous_year": 0.0},
  "total_income": {"current_year": 416194.0, "previous_year": 0.0},
  "employee_benefit_expense": {"current_year": 129820.0, "previous_year": 0.0},
  "payment_to_auditors": {"current_year": 10000.0, "previous_year": 7500.0},
  "power_and_fuel": {"current_year": 0.0, "previous_year": 0.0},
  "finance_costs": {"current_year": 5256.0, "previous_year": 1250.0},
  "depreciation_and_amortisation": {"current_year": 26659.0, "previous_year": 0.0},
  "other_expenses": {"current_year": 273098.0, "previous_year": 19500.0},
  "total_expenses": {"current_year": 444833.0, "previous_year": 28250.0},
  "profit_before_exceptional_items": {"current_year": -28640.0, "previous_year": -28250.0},
  "profit_before_tax": {"current_year": -28640.0, "previous_year": -28250.0},
  "deferred_tax": {"current_year": 12200.0, "previous_year": -3926.0},
  "tax_expense": {"current_year": 12200.0, "previous_year": -3926.0},
  "profit_after_tax": {"current_year": -40840.0, "previous_year": -24324.0},
  "earnings_per_share_basic": {"current_year": -4.08, "previous_year": -5.53},
  "gross_transaction_as_18": {"current_year": 1221644.0, "previous_year": 815620.0},
  "param_rent_paid": {"current_year": 245818.0, "previous_year": 0.0},
  "pcs_code": {"current_year": 9996.0, "previous_year": 0.0},
  "pcs_description": "Sports activities and recreational services",
  "pcs_turnover": {"current_year": 416194.0, "previous_year": 0.0},
  "pcs_highest_code": {"current_year": 99965900.0, "previous_year": 0.0},
  "pcs_highest_description": "Sports activities and recreational services",
  "pcs_highest_turnover": {"current_year": 416194.0, "previous_year": 0.0},
  "net_worth": {"current_year": -55164.0, "previous_year": -14323.0}
}
'''

data = json.loads(json_str)

print("=" * 135)
print(f"{'Field Label':<42} | {'Extracted CY':<18} | {'CS Ref CY':<18} | {'Match Status':<22} | {'Extracted PY':<16}")
print("=" * 135)

matches_cy = 0
matches_scaled = 0
total_cy = 0

for row, col_cy, col_py, key, label in field_mappings:
    ref_cy = sheet.cell(row=row, column=col_cy).value if col_cy else None
    ref_py = sheet.cell(row=row, column=col_py).value if col_py else None
    
    val = data.get(key)
    if isinstance(val, dict):
        ext_cy = val.get("current_year")
        ext_py = val.get("previous_year")
    else:
        ext_cy = val
        ext_py = None
        
    ref_cy_str = str(ref_cy).strip() if ref_cy is not None else "EMPTY"
    ref_py_str = str(ref_py).strip() if ref_py is not None else "EMPTY"
    
    ext_cy_str = str(ext_cy) if ext_cy is not None else "EMPTY"
    ext_py_str = str(ext_py) if ext_py is not None else "EMPTY"
    
    # Check match for CY
    match_status = "❌ NO"
    try:
        if ext_cy is not None and ref_cy is not None:
            if str(ext_cy_str).strip() == str(ref_cy_str).strip():
                match_status = "✅ YES (Exact)"
            elif abs(float(ext_cy) - float(ref_cy)) < 2.0:
                match_status = "✅ YES (~Rupee)"
            elif abs((float(ext_cy) / 100.0) - float(ref_cy)) < 2.0:
                match_status = "⚠️ YES (if / 100)"
            elif abs((float(ext_cy) / 1000.0) - float(ref_cy)) < 2.0:
                match_status = "⚠️ YES (if / 1000)"
    except:
        if ext_cy_str.lower() in ref_cy_str.lower() or ref_cy_str.lower() in ext_cy_str.lower():
            match_status = "✅ YES (Text)"
            
    if ref_cy_str != "EMPTY":
        total_cy += 1
        if "✅" in match_status or "⚠️" in match_status:
            matches_cy += 1
            
    print(f"{label:<42} | {ext_cy_str:<18} | {ref_cy_str:<18} | {match_status:<22} | {ext_py_str:<16}")

wb.close()
print("=" * 130)
print(f"VERIFICATION SCORE: {matches_cy} / {total_cy} fields matched CS reference!")
print("=" * 130)
