"""
Compare the CS-completed reference Excel with our auto-filled Excel.
Scrapes both files and outputs a detailed diff of every cell that differs.
"""
import openpyxl

ref_path = r"C:\Users\RIYAS\Downloads\Copy of AOC-4_U92410KL2020PTC065216_2021-2022_20260728.xlsx"
our_path = r"C:\Users\RIYAS\Downloads\AOC-4_U92410KL2020PTC065216_2021-2022_20260729_FILLED.xlsx"

def load_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb["FORM"], wb

ref_sheet, ref_wb = load_sheet(ref_path)
our_sheet, our_wb = load_sheet(our_path)

print("=" * 120)
print("COMPREHENSIVE CELL-BY-CELL COMPARISON")
print(f"Reference (CS-completed): {ref_path}")
print(f"Our auto-filled:          {our_path}")
print("=" * 120)

# Scan all rows from 1 to 600, columns A(1) to P(16)
diffs = []
for row in range(1, 601):
    for col in range(1, 17):
        ref_val = ref_sheet.cell(row=row, column=col).value
        our_val = our_sheet.cell(row=row, column=col).value
        
        # Normalize
        ref_str = str(ref_val).strip() if ref_val is not None else ""
        our_str = str(our_val).strip() if our_val is not None else ""
        
        # Skip if both empty
        if not ref_str and not our_str:
            continue
        
        # Skip if identical
        if ref_str == our_str:
            continue
        
        # Try numeric comparison (handle float precision)
        try:
            ref_num = float(ref_str) if ref_str else None
            our_num = float(our_str) if our_str else None
            if ref_num is not None and our_num is not None and abs(ref_num - our_num) < 0.01:
                continue
        except (ValueError, TypeError):
            pass
        
        col_letter = chr(64 + col) if col <= 26 else f"C{col}"
        
        # Get row label for context
        label_b = str(ref_sheet.cell(row=row, column=2).value or "").strip()
        label_c = str(ref_sheet.cell(row=row, column=3).value or "").strip()
        label = label_c if label_c else label_b
        
        diffs.append({
            "row": row,
            "col": col,
            "col_letter": col_letter,
            "label": label[:60],
            "ref_val": ref_str[:50],
            "our_val": our_str[:50],
        })

# Print all diffs grouped by section
print(f"\nFound {len(diffs)} differences:\n")
print(f"{'Row':<5} {'Col':<4} {'Label':<62} {'REF (CS-completed)':<30} {'OURS (auto-filled)':<30}")
print("-" * 135)

current_section = ""
for d in diffs:
    row = d["row"]
    # Add section headers
    if row < 200:
        section = "GENERAL INFO"
    elif row < 247:
        section = "PART I - BALANCE SHEET"
    elif row < 329:
        section = "II BREAK-UP OF BALANCE SHEET"
    elif row < 379:
        section = "III FINANCIAL PARAMS - BS"
    elif row < 440:
        section = "IV SHARE CAPITAL & V SBN & VI COST"
    elif row < 497:
        section = "SEGMENT II - P&L STATEMENT"
    elif row < 523:
        section = "II DETAILED P&L - FX"
    elif row < 538:
        section = "III FINANCIAL PARAMS - PL"
    elif row < 560:
        section = "IV PRINCIPAL PRODUCTS"
    else:
        section = "SEGMENT III+ (RPT, AUDITOR, CSR, MISC)"
    
    if section != current_section:
        print(f"\n>>> {section} <<<")
        current_section = section
    
    print(f"{d['row']:<5} {d['col_letter']:<4} {d['label']:<62} {d['ref_val']:<30} {d['our_val']:<30}")

ref_wb.close()
our_wb.close()

print(f"\n{'=' * 120}")
print(f"SUMMARY: {len(diffs)} total differences found.")
print(f"{'=' * 120}")
