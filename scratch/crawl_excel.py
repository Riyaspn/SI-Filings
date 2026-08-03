import openpyxl

def inspect_excel(filepath):
    print(f"Inspecting: {filepath}")
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"Failed to load workbook. Ensure it is not currently open in Excel. Error: {e}")
        return

    sheet_name = "FORM"
    if sheet_name in wb.sheetnames:
        print(f"\n--- DUMPING {sheet_name} ---")
        sheet = wb[sheet_name]
        
        # Print the first 500 rows, columns A to K
        for row in range(1, 501):
            row_data = []
            for col in range(1, 12): # A to K
                cell = sheet.cell(row=row, column=col)
                val = cell.value
                if val is not None and str(val).strip() != "":
                    # Clean up formatting for display
                    val = str(val).strip().replace("\n", " ")
                    row_data.append(f"{cell.coordinate}: {val}")
            
            if row_data:
                print(" | ".join(row_data))
    else:
        print(f"Could not find sheet '{sheet_name}'. Available: {wb.sheetnames}")

if __name__ == "__main__":
    filepath = r"C:\RIYAS\Sharp INtell\SI Filings\AOC-4_U92410KL2020PTC065216_2021-2022_20260729.xlsx"
    inspect_excel(filepath)
