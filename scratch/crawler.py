import openpyxl
import json
import os
import re

def create_gap_analysis():
    # 1. Load JSON data (Verification page)
    json_path = r"c:\RIYAS\Sharp INtell\SI Filings\aoc4_diagnosis.json"
    print(f"Loading extracted data from: {json_path}")
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            diagnosis_data = json.load(f).get("data", {})
    except Exception as e:
        print(f"Error loading JSON: {e}")
        return

    # 2. Load the specific filled Excel file from Downloads
    excel_path = r"C:\Users\RIYAS\Downloads\AOC-4_U92410KL2020PTC065216_2021-2022_20260729_FILLED.xlsx"
    print(f"Loading Excel file: {excel_path}")
    
    if not os.path.exists(excel_path):
        print(f"File not found: {excel_path}")
        return
        
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb["FORM"]
    except Exception as e:
        print(f"Error loading Excel: {e}")
        return

    # Mapped rows (Current working mapping in excel_populator.py)
    # Using a reverse map to easily check if a row is currently mapped
    current_mapping = {
        204: "share_capital", 205: "reserves_and_surplus", 206: "money_received_share_warrants",
        208: "share_application_money", 212: "long_term_borrowings", 213: "deferred_tax_liabilities",
        214: "other_long_term_liabilities", 215: "long_term_provisions", 218: "short_term_borrowings",
        220: "trade_payables_msme", 221: "trade_payables_others", 222: "other_current_liabilities",
        223: "short_term_provisions", 229: "tangible_assets", 231: "intangible_assets",
        232: "capital_wip", 233: "intangible_assets_under_dev", 234: "non_current_investments",
        235: "deferred_tax_assets", 236: "long_term_loans_advances", 237: "other_non_current_assets",
        239: "current_investments", 240: "inventories", 241: "trade_receivables", 242: "cash_and_bank_balances",
        243: "short_term_loans_advances", 244: "other_current_assets", 449: "rev_sale_goods_mfg",
        450: "rev_sale_goods_traded", 451: "rev_sale_services", 457: "oi_dividend", 458: "oi_interest",
        459: "oi_net_gain_investments", 460: "oi_other_non_operating", 463: "cost_of_materials_consumed",
        464: "purchases_of_stock_in_trade", 465: "changes_in_inventories", 468: "employee_benefit_expense",
        469: "managerial_remuneration", 470: "payment_to_auditors", 471: "insurance_expenses",
        472: "power_and_fuel", 473: "finance_costs", 474: "depreciation_and_amortisation",
        475: "other_expenses", 478: "exceptional_items", 480: "extraordinary_items",
        483: "current_tax", 484: "deferred_tax"
    }

    # Generate markdown report
    output_md = r"c:\Users\RIYAS\.gemini\antigravity-ide\brain\2795f6fe-c257-4886-8374-88a4e78e10dd\gap_analysis.md"
    
    with open(output_md, "w", encoding="utf-8") as out:
        out.write("# AOC-4 Excel vs JSON Gap Analysis\n\n")
        out.write("| Row | Excel Field Description | CY Col | PY Col | Mapped in Script? | In JSON App Data? | Notes |\n")
        out.write("|---|---|---|---|---|---|---|\n")

        # Scan rows 190 to 600 which contain the financial tables
        for row in range(190, 600):
            # Look for labels in Col B or C
            label = str(sheet.cell(row=row, column=3).value or "").strip()
            if not label or label.lower() in ("none", ""):
                label = str(sheet.cell(row=row, column=2).value or "").strip()
                if not label or label.lower() in ("none", ""):
                    continue

            # Skip header lines that don't need numbers
            if len(label) < 3 or label.startswith("PART") or label.startswith("SEGMENT"):
                continue

            # Find where the data columns are. Mostly G/J, but some are N/O or H/J
            # Let's just check G, J, N, O for numbers to identify the data columns
            cols_to_check = [7, 10, 14, 15] # G, J, N, O
            cy_col_used = ""
            py_col_used = ""
            
            for c in cols_to_check:
                cell_val = sheet.cell(row=row, column=c).value
                # If there's a 0 or some number in the filled excel, it's an input field
                if cell_val is not None and str(cell_val).strip() != "":
                    if c in (7, 14): cy_col_used = openpyxl.utils.get_column_letter(c)
                    if c in (10, 15): py_col_used = openpyxl.utils.get_column_letter(c)

            # If this row is mapped in our populator
            mapped_key = current_mapping.get(row, "NO")
            
            # Is it in the JSON data?
            in_json = "NO"
            if mapped_key != "NO":
                if mapped_key in diagnosis_data:
                    in_json = "YES"

            # Clean label for markdown table
            clean_label = label.replace("\n", " ").replace("|", "-")
            if len(clean_label) > 65:
                clean_label = clean_label[:62] + "..."

            out.write(f"| {row} | {clean_label} | {cy_col_used} | {py_col_used} | {mapped_key} | {in_json} | |\n")
            
    print(f"\nGap Analysis complete! Report saved to:\n{output_md}")

if __name__ == "__main__":
    create_gap_analysis()
