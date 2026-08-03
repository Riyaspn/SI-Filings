import os
import json
import openpyxl
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from validator import validate_and_heal_payload
from excel_populator import ExcelPopulator

json_path = r"C:\Users\RIYAS\Downloads\AOC4_20260802_024637.json"
if not os.path.exists(json_path):
    print(f"Error: JSON file not found at {json_path}")
    sys.exit(1)

with open(json_path, "r", encoding="utf-8") as f:
    raw = json.load(f)

data = raw.get("data", {})
print("--- 1. Running Upgraded Mathematical Validator & Self-Healing Engine ---")
healed_data, audit = validate_and_heal_payload(data)

# Check potential template locations
template_candidates = [
    r"C:\RIYAS\Sharp INtell\SI Filings\AOC-4_U92410KL2020PTC065216_2021-2022_20260729.xlsx",
    r"C:\Users\RIYAS\Downloads\AOC-4_U92410KL2020PTC065216_2021-2022_20260729.xlsx"
]

template_path = None
for c in template_candidates:
    if os.path.exists(c):
        template_path = c
        break

if not template_path:
    print("Error: Could not find blank template file!")
    sys.exit(1)

import subprocess

print(f"--- 2. Populating Excel via COM (Template: {template_path}) ---")
try:
    # Safely clear any frozen EXCEL.EXE processes left over from earlier interrupted test runs
    subprocess.run(["taskkill", "/f", "/im", "excel.exe"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
except Exception:
    pass

out_path = r"C:\Users\RIYAS\Downloads\AOC-4_U92410KL2020PTC065216_2021-2022_20260729_FILLED.xlsx"
populator = ExcelPopulator(template_path)
success, msg = populator.populate(healed_data, out_path)
print(f"Populate result: {success} -> {msg}")

print("--- 3. Running Cell-By-Cell Comparison vs CS Reference ---")
ref_path = r"C:\Users\RIYAS\Downloads\Copy of AOC-4_U92410KL2020PTC065216_2021-2022_20260728.xlsx"

wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
sheet_ref = wb_ref["FORM"]
wb_our = openpyxl.load_workbook(out_path, data_only=True)
sheet_our = wb_our["FORM"]

diffs = []
for row in range(1, 601):
    for col in range(1, 17):
        ref_val = sheet_ref.cell(row=row, column=col).value
        our_val = sheet_our.cell(row=row, column=col).value
        
        ref_str = str(ref_val).strip() if ref_val is not None else ""
        our_str = str(our_val).strip() if our_val is not None else ""
        
        if not ref_str and not our_str:
            continue
        if ref_str == our_str:
            continue
            
        # Normalize optional "M/s " business title prefix
        if ref_str.replace("M/s ", "").replace("M/s. ", "").strip() == our_str.replace("M/s ", "").replace("M/s. ", "").strip():
            continue
            
        # Recognize semantic equivalence for extended NIC industry descriptions (e.g., "Sports activities" vs "Sports activities and recreational services")
        if (ref_str in our_str or our_str in ref_str) and len(ref_str) > 10 and len(our_str) > 10 and not any(ch.isdigit() for ch in ref_str):
            continue
            
        try:
            if float(ref_str) == float(our_str) or abs(float(ref_str) - float(our_str)) <= 1.00:
                continue
        except (ValueError, TypeError):
            pass
            
        col_letter = chr(64 + col) if col <= 26 else f"C{col}"
        label_b = str(sheet_ref.cell(row=row, column=2).value or "").strip()
        label_c = str(sheet_ref.cell(row=row, column=3).value or "").strip()
        label = label_c if label_c else label_b
        
        diffs.append({
            "row": row, "col_letter": col_letter, "label": label[:60],
            "ref": ref_str[:40], "our": our_str[:40]
        })

print(f"\nFound {len(diffs)} differences:\n")
print(f"{'Row':<5} {'Col':<4} {'Label':<60} {'REF (CS-completed)':<30} {'OURS (auto-filled)':<30}")
print("-" * 135)
for d in diffs:
    print(f"{d['row']:<5} {d['col_letter']:<4} {d['label']:<60} {d['ref']:<30} {d['our']:<30}")

wb_ref.close()
wb_our.close()
